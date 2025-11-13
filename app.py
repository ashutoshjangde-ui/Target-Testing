# app_masterfile.py

import io
import json
import os
import re
import sqlite3
from difflib import SequenceMatcher
from textwrap import dedent

import pandas as pd
import streamlit as st
from openpyxl import load_workbook

st.set_page_config(page_title="Masterfile Automation", page_icon="📦", layout="wide")

# =========================
# Deterministic mapping store (SQLite)
# =========================
DB_PATH = ".data/mappings.db"
SIM_THRESHOLD = 0.95  # high threshold for sensitive data

def _db():
    os.makedirs(os.path.dirname(DB_PATH), exist_ok=True)
    return sqlite3.connect(DB_PATH)

def init_db():
    with _db() as c:
        c.execute("""
        CREATE TABLE IF NOT EXISTS maps(
            vendor TEXT,
            version INTEGER,
            master_norm TEXT,
            alias TEXT,
            PRIMARY KEY(vendor, version, master_norm)
        )
        """)
        c.execute("""
        CREATE TABLE IF NOT EXISTS meta(
            vendor TEXT PRIMARY KEY,
            latest_version INTEGER
        )
        """)

def get_latest_version(vendor: str) -> int:
    with _db() as c:
        row = c.execute("SELECT latest_version FROM meta WHERE vendor=?", (vendor,)).fetchone()
        return row[0] if row else 0

def load_vendor_map(vendor: str, version: int | None = None) -> tuple[dict, int]:
    """Return ({master_norm -> alias}, used_version)"""
    v = get_latest_version(vendor) if version is None else version
    if v == 0:
        return {}, 0
    with _db() as c:
        rows = c.execute("SELECT master_norm, alias FROM maps WHERE vendor=? AND version=?", (vendor, v)).fetchall()
    return ({mn: al for mn, al in rows}, v)

def save_vendor_map(vendor: str, mapping: dict, base_version: int = 0) -> int:
    """mapping: {master_norm -> alias}"""
    new_v = (base_version or get_latest_version(vendor)) + 1
    with _db() as c:
        for mn, al in mapping.items():
            c.execute("REPLACE INTO maps(vendor, version, master_norm, alias) VALUES(?,?,?,?)",
                      (vendor, new_v, mn, al))
        c.execute("REPLACE INTO meta(vendor, latest_version) VALUES(?,?)", (vendor, new_v))
    return new_v

# =========================
# Helpers
# =========================
def norm(s: str) -> str:
    """Normalize header strings for robust matching."""
    if s is None:
        return ""
    x = str(s).strip().lower()
    x = re.sub(r"\s*-\s*en\s*[-_ ]\s*us\s*$", "", x)  # strip "- en-US"
    x = x.replace("–", "-").replace("—", "-").replace("−", "-")
    x = re.sub(r"[._/\\-]+", " ", x)
    x = re.sub(r"[^0-9a-z\s]+", " ", x)
    return re.sub(r"\s+", " ", x).strip()

def sim(a: str, b: str) -> float:
    return SequenceMatcher(None, norm(a), norm(b)).ratio()

def top_matches(query, candidates, k=3):
    q = norm(query)
    scored = [(SequenceMatcher(None, q, norm(c)).ratio(), c) for c in candidates]
    scored.sort(key=lambda t: t[0], reverse=True)
    return scored[:k]

def worksheet_used_cols(ws, header_rows=(1,), hard_cap=512, empty_streak_stop=8):
    max_try = min(ws.max_column, hard_cap)
    last_nonempty, streak = 0, 0
    for c in range(1, max_try + 1):
        any_val = False
        for r in header_rows:
            v = ws.cell(row=r, column=c).value
            if v not in (None, ""):
                any_val = True
                break
        if any_val:
            last_nonempty = c
            streak = 0
        else:
            streak += 1
            if streak >= empty_streak_stop:
                break
    return max(last_nonempty, 1)

def nonempty_rows(df: pd.DataFrame) -> int:
    if df.empty:
        return 0
    tmp = df.replace("", pd.NA)
    return tmp.dropna(how="all").shape[0]

def load_canon_synonyms() -> dict:
    """
    Optional local file 'canon.json':
    {
      "Product Title": ["Item Name","Title"],
      "Partner SKU": ["Seller SKU","SKU","item_sku"]
    }
    Keys are canonical/master display headers; values are alias lists.
    """
    path = "canon.json"
    if os.path.exists(path):
        try:
            with open(path, "r", encoding="utf-8") as f:
                raw = json.load(f)
                # ensure lists
                canon = {k: (v if isinstance(v, list) else [v]) for k, v in raw.items()}
                return canon
        except Exception:
            pass
    return {}

def build_aliases_for_master(master_displays, vendor_map, canon_synonyms):
    """
    Return dict: { master_norm -> [aliases in priority order] }
    Priority: saved vendor alias -> canon synonyms -> master display itself
    """
    out = {}
    for disp in master_displays:
        mn = norm(disp)
        if not mn:
            continue
        aliases = []
        # 1) saved alias for this canonical master
        saved_alias = vendor_map.get(mn)
        if saved_alias:
            aliases.append(saved_alias)
        # 2) canon synonyms by key equal to master display (case-insensitive)
        syn = canon_synonyms.get(disp, []) or canon_synonyms.get(disp.strip(), [])
        # also allow matching by normalized key
        if not syn:
            # try to find a canon key with same normalized text
            for k in canon_synonyms.keys():
                if norm(k) == mn:
                    syn = canon_synonyms.get(k, [])
                    break
        # de-dup while preserving order
        seen = set(a.lower().strip() for a in aliases)
        for a in syn:
            al = str(a)
            if al.lower().strip() not in seen:
                aliases.append(al)
                seen.add(al.lower().strip())
        # 3) fallback to display itself
        if disp not in seen:
            aliases.append(disp)
        out[mn] = aliases
    return out

def pick_best_onboarding_sheet(uploaded_file, mapping_aliases_by_master):
    """
    Same scoring as before, now using auto-built alias lists.
    """
    uploaded_file.seek(0)
    xl = pd.ExcelFile(uploaded_file)

    best = None
    best_score = -1
    best_info = ""
    for sheet in xl.sheet_names:
        try:
            df = xl.parse(sheet_name=sheet, header=0, dtype=str)
            df = df.fillna("")
            df.columns = [str(c).strip() for c in df.columns]
        except Exception:
            continue

        header_set = {norm(c) for c in df.columns}
        matches = 0
        for master_norm, aliases in mapping_aliases_by_master.items():
            if any(norm(a) in header_set for a in aliases):
                matches += 1

        rows = nonempty_rows(df)
        score = matches + min(rows, 1) * 0.01
        if score > best_score:
            best = (df, sheet)
            best_score = score
            best_info = f"matched headers: {matches}, non-empty rows: {rows}"

    if best is None:
        raise ValueError("No readable sheet found in onboarding workbook.")

    df, sheet_name = best
    return df, sheet_name, best_info

# Unique sentinel for the special "Listing Action" fill
SENTINEL_LISTING_ACTION = object()

# =========================
# UI
# =========================
st.title("📦 Masterfile Automation")
st.caption("Map onboarding columns to master template headers and generate a ready-to-upload masterfile.")

with st.expander("ℹ️ Instructions", expanded=True):
    instructions = dedent("""
    - **Masterfile template (.xlsx)**  
      - Row **1** = display labels  
      - Row **2** = internal keys/helper labels  
      - Data is written starting at **Row 3** (this tool preserves template styles).

    - **Onboarding sheet (.xlsx)**  
      - Row **1** = **headers**  
      - Row **2+** = data

    - **No mapping file needed**  
      - The app auto-maps using your **saved per-vendor mapping** (versioned), optional **canon.json** synonyms, and strict similarity checks.  
      - Any new/unknown headers must be reviewed in the log; unmatched master columns stay blank (never guessed below the high threshold).
    """)
    st.markdown(instructions)

st.divider()

c1, c2, c3 = st.columns([1, 1, 1])
with c1:
    vendor_id = st.text_input("🏷️ Vendor ID (required)")
with c2:
    masterfile_file = st.file_uploader("📄 Upload Masterfile Template (.xlsx)", type=["xlsx"])
with c3:
    onboarding_file = st.file_uploader("🧾 Upload Onboarding Sheet (.xlsx)", type=["xlsx"])

st.divider()
go = st.button("🚀 Generate Final Masterfile", type="primary")

log_area = st.container()
download_area = st.container()

# =========================
# Main Action
# =========================
if go:
    init_db()

    with log_area:
        st.markdown("### 📝 Log")
        log = st.empty()

        def slog(msg):
            log.markdown(msg)

        # Validate inputs
        if not vendor_id.strip():
            st.error("Please enter a **Vendor ID**.")
            st.stop()
        if not masterfile_file or not onboarding_file:
            st.error("Please upload both **Masterfile Template** and **Onboarding Sheet**.")
            st.stop()

        slog("⏳ Reading masterfile template…")
        try:
            master_wb = load_workbook(masterfile_file, keep_links=False)
            master_ws = master_wb.active
        except Exception as e:
            st.error(f"Could not read **Masterfile**: {e}")
            st.stop()

        used_cols = worksheet_used_cols(master_ws, header_rows=(1, 2))
        master_displays = [master_ws.cell(row=1, column=c).value or "" for c in range(1, used_cols + 1)]

        # Load vendor's last mapping + optional canon synonyms
        vendor_map, v_used = load_vendor_map(vendor_id.strip())
        canon_synonyms = load_canon_synonyms()
        mapping_aliases_by_master = build_aliases_for_master(master_displays, vendor_map, canon_synonyms)

        # --- pick best onboarding sheet automatically ---
        slog("🔎 Selecting best onboarding sheet…")
        try:
            best_df, best_sheet, info = pick_best_onboarding_sheet(onboarding_file, mapping_aliases_by_master)
            st.success(f"Using onboarding sheet: **{best_sheet}** ({info})")
        except Exception as e:
            st.error(f"Could not find a suitable onboarding sheet: {e}")
            st.stop()

        on_df = best_df.fillna("")
        on_headers = list(on_df.columns)
        series_by_alias = {norm(h): on_df[h] for h in on_headers}

        master_to_source = {}   # column index -> (Series) or SENTINEL_LISTING_ACTION
        chosen_alias = {}       # column index -> alias actually used (for reporting)
        unmatched = []
        report_lines = []

        report_lines.append("#### 🔗 Mapping Summary (saved > canon > exact > high-threshold sim)")
        for c, m_disp in enumerate(master_displays, start=1):
            disp_norm = norm(m_disp)
            if not disp_norm:
                continue

            aliases = mapping_aliases_by_master.get(disp_norm, [m_disp])
            resolved_series = None
            resolved_alias = None

            # 1) saved + canon + display itself via exact lookup
            for a in aliases:
                a_norm = norm(a)
                if a_norm in series_by_alias:
                    resolved_series = series_by_alias[a_norm]
                    resolved_alias = a
                    break

            # 2) high-threshold similarity fallback (suggestion becomes selection only if ≥ SIM_THRESHOLD)
            if resolved_series is None:
                candidates = [(sim(m_disp, h), h) for h in on_headers]
                candidates.sort(reverse=True)
                if candidates and candidates[0][0] >= SIM_THRESHOLD:
                    resolved_series = series_by_alias[norm(candidates[0][1])]
                    resolved_alias = candidates[0][1]

            if resolved_series is not None:
                master_to_source[c] = resolved_series
                chosen_alias[c] = resolved_alias
                source_kind = "saved/canon/exact" if resolved_alias in aliases else f"similarity {round(sim(m_disp, resolved_alias)*100,1)}%"
                report_lines.append(f"- ✅ **{m_disp}** ← `{resolved_alias}` ({source_kind})")
            else:
                if disp_norm == norm("Listing Action (List or Unlist)"):
                    master_to_source[c] = SENTINEL_LISTING_ACTION
                    report_lines.append(f"- 🟨 **{m_disp}** ← (will fill `'List'`)")
                else:
                    unmatched.append(m_disp)
                    suggestions = top_matches(m_disp, on_headers, 3)
                    sug_txt = ", ".join(f"`{name}` ({round(sc*100,1)}%)" for sc, name in suggestions) if suggestions else "*none*"
                    report_lines.append(f"- ❌ **{m_disp}** ← *no match*. Suggestions: {sug_txt}")

        st.markdown("\n".join(report_lines))

        # Safety: block export if nothing mapped (avoids blank files on sensitive data)
        mapped_cols = [c for c in master_to_source.keys() if master_to_source[c] is not None]
        if not mapped_cols:
            st.error("No columns were confidently mapped. Please adjust the onboarding headers or add `canon.json` synonyms.")
            st.stop()

        # Write values to master starting row 3
        slog("🛠️ Writing data…")
        out_row = 3
        num_rows = len(on_df)

        for i in range(num_rows):
            for c in range(1, used_cols + 1):
                src = master_to_source.get(c, None)
                if src is None:
                    continue
                if src is SENTINEL_LISTING_ACTION:
                    master_ws.cell(row=out_row + i, column=c, value="List")
                elif isinstance(src, pd.Series):
                    if i < len(src):
                        master_ws.cell(row=out_row + i, column=c, value=str(src.iloc[i]))

        # Persist mapping for this vendor (only what we actually used)
        slog("💾 Saving mapping…")
        # Build {master_norm -> chosen_alias_text}
        final_map = {}
        for c, m_disp in enumerate(master_displays, start=1):
            if c in chosen_alias:
                final_map[norm(m_disp)] = str(chosen_alias[c]).strip()

        new_version = save_vendor_map(vendor_id.strip(), final_map, base_version=v_used)

        # Save workbook to buffer
        slog("📦 Finalizing workbook…")
        bio = io.BytesIO()
        master_wb.save(bio)
        bio.seek(0)

        with download_area:
            st.success(f"✅ Final masterfile is ready! (Saved mapping v{new_version} for vendor **{vendor_id}**)")

            st.download_button(
                "⬇️ Download Final Masterfile",
                data=bio.getvalue(),
                file_name="final_masterfile_real.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

            if unmatched:
                st.info(
                    "Some master columns had no match and were left blank:\n\n- " +
                    "\n- ".join(unmatched)
                )
